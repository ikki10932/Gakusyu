import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_styled_excel():
    wb = openpyxl.Workbook()
    
    # 共通スタイル設定
    header_font = Font(name='Meiryo UI', bold=True, color="FFFFFF")
    body_font = Font(name='Meiryo UI')
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_header_style(sheet, columns):
        for col_idx, (name, width) in enumerate(columns, 1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        sheet.freeze_panes = "A4"

    # Sheet 1: 問題点一覧
    ws1 = wb.active
    ws1.title = "問題点一覧"
    cols1 = [("カテゴリ", 25), ("問題点", 45), ("詳細・リスク", 65), ("改善の方向性", 65)]
    apply_header_style(ws1, cols1)
    
    data1 = [
        ("情報の混在", "目的、コマンド、結果が混在している", "結論が即座に判別しにくく、誤解を招くリスクがある", "「知見の共有」と「手法の標準化」にシフトし、構成を整理する"),
        ("コンテキストの依存", "特定のAIモデルの挙動記述が中心", "数年後に参照した際の有用性が低下する懸念がある", "モデル依存の記述を減らし、汎用的なワークフローを提示する"),
        ("リンクの不透明性", "多数のローカルパスが説明なしに記載", "環境変化時に追跡不能になるリスクがある", "ファイルごとの役割説明を加え、相対的な関係を明示する")
    ]
    for r_idx, row in enumerate(data1, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = top_left_alignment if c_idx > 1 else center_alignment
            cell.border = border

    # Sheet 2: 推奨構成定義
    ws2 = wb.create_sheet("推奨構成定義")
    cols2 = [("章題", 35), ("具体的説明", 55), ("必須記載項目", 65), ("補強提案", 65)]
    apply_header_style(ws2, cols2)
    
    data2 = [
        ("検証環境と使用ツール", "Antigravityの構成やSKILL.mdの役割", "バージョン情報、SKILL定義の意図", "SKILL.mdのバージョン管理方針の提示"),
        ("標準解析ワークフロー", "モデルに依存しない汎用的な手順", "入力データから出力を得る論理ステップ", "エラーハンドリング集のリスト化"),
        ("検証結果と考察", "得られた成果物例と特性分析", "生成された成果物(xlsx/md)の評価", "AI未使用時との工数比較データの追加")
    ]
    for r_idx, row in enumerate(data2, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = top_left_alignment
            cell.border = border

    # Sheet 3: 欠落情報・要対応
    ws3 = wb.create_sheet("欠落情報・要対応")
    cols3 = [("欠落している情報", 55), ("理由・背景", 65), ("対応優先度", 15)]
    apply_header_style(ws3, cols3)
    
    data3 = [
        ("核となる指示内容(rules.md/Instructions.md)", "どのようなポリシーで解析させているか不明なため", "高"),
        ("AI未使用時との比較データ", "定量的な効果測定ができていないため", "中")
    ]
    for r_idx, row in enumerate(data3, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = center_alignment if c_idx == 3 else top_left_alignment
            cell.border = border

    save_path = r"d:\_Develop\Skills\MySkills\review\output\2026-02-27_資料設計レビューアドバイス.xlsx"
    wb.save(save_path)
    print(f"Excel created at: {save_path}")

if __name__ == "__main__":
    create_styled_excel()
