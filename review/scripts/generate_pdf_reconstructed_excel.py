import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_reconstructed_excel():
    wb = openpyxl.Workbook()
    
    header_font = Font(name='Meiryo UI', bold=True, color="FFFFFF")
    body_font = Font(name='Meiryo UI')
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border_side = Side(style='thin')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    def apply_header_style(sheet, columns):
        for col_idx, (name, width) in enumerate(columns, 1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        sheet.freeze_panes = "A4"

    # Sheet 1: AI検証レポート（再構成版）
    ws1 = wb.active
    ws1.title = "AI検証レポート（再構成版）"
    ws1.cell(row=1, column=1, value="AIエージェントによるソースコード解析と資料生成の有効性検証レポート").font = Font(name='Meiryo UI', bold=True, size=14)
    ws1.row_dimensions[1].height = 25
    
    cols1 = [("構成（章題）", 30), ("記載事項", 30), ("具体的な内容（手順・結果・設定など）", 80)]
    apply_header_style(ws1, cols1)
    
    data1 = [
        ("1. 検証の背景と目的", "背景", "既存の資料の構造や表現が作成者に依存（属人化）しており、長期的に第三者が理解可能な設計が整っていなかった。"),
        ("1. 検証の背景と目的", "目的", "AIにソースコード解析を行わせることで「解析手順」や「生成資料」の品質を標準化し、属人化の解消と長期理解可能な資料設計を定着させること。"),
        ("2. 検証環境とツールスタック", "使用ツール", "AIツール: Antigravity\n使用モデル: Gemini 3.1 pro, Gemini Flash, Claude sonnet"),
        ("2. 検証環境とツールスタック", "使用ルール・SKILL等の役割定義", "・rules.md: 許可不要の操作範囲明示、出力へのメタ情報付与、ファイル名上書き禁止など\n・Instructions.md: 手順命令\n・Excel_Layout_Skill.md: エクセル出力時のレイアウト指定\n・VC6MFC_skill.md: VC6MFCの解析指定"),
        ("3. 標準解析ワークフロー", "Step 1: Input資料の投入", "対象とするソースファイル(.cpp等)および設定ファイル(.dat)と、上記の定義ファイル(rules.md, Instructions.md等)をAIセッションに都度投入する。"),
        ("3. 標準解析ワークフロー", "Step 2: プロンプト（指示）の実行", "明確で省略を許さないプロンプトを実行する。\n例：「router01W.datに書き込まれている内容の1行毎に、どの関数でSDkRouterのどのメンバ変数を用いて書き込まれたか解析してレポートを出力してください。省略はしないでください」"),
        ("3. 標準解析ワークフロー", "Step 3: 結果の出力と整形", "Markdown形式のレポートや、指定レイアウト（Excel）にて最終出力として保存させる。"),
        ("4. 検証結果と得られたベストプラクティス", "成果物例", "ルータ設定ファイルの入出力元解析レポート（xlsxファイルやmdファイル）の生成に成功した。\n出力先例: D:\_Develop\Skills\MySkills\router\output\2026-02-26_dat_io_report.xlsx"),
        ("4. 検証結果と得られたベストプラクティス", "ベストプラクティス", "「省略はしないでください」と明確に指示することで、途中で推論や処理が打ち切られるのを防ぐことができる。"),
        ("5. 発生した課題と対処法", "モデルによる出力のばらつき", "モデル（例: Gemini 3.1 pro）によっては、表の列幅設定が無視されたり、関数表記が雑になる場合がある。\n対処: 別のモデル(Claude sonnet等)での処理や、Pythonスクリプトを用いたxlsx生成など、出力手法の切り替えを行う。"),
        ("5. 発生した課題と対処法", "タイムアウトやエラーの頻発", "Loading→working→Generatingの処理がループして時間がかかったり、「Agent terminated due to error」が頻発することがある。\n対処: エラー時はリトライ（Retry）を実行する。将来的には、バッチ化によりエラー時の自動再試行手順を標準化することが望ましい。"),
        ("6. リファレンスと今後の展開", "コンテキスト情報・参照パス", "入力元ベース: D:/_Develop/Skills/MySkills/router/.agent/\n出力先ベース: D:\_Develop\Skills\MySkills\router\output\\"),
        ("6. リファレンスと今後の展開", "今後の展開・方針", "・各種プログラム(PG)への応用\n・AI提案の精度向上に向けた skill.md 自体の改良\n・バッチ処理によるレビューの自動化\n・段階的な運用導入による「工数削減」と「品質向上」の両立")
    ]
    
    for r_idx, row in enumerate(data1, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = top_left_alignment
            cell.border = border

    output_dir = r"d:\_Develop\Skills\MySkills\review\output"
    os.makedirs(output_dir, exist_ok=True)
    save_path = os.path.join(output_dir, "2026-03-02_1pdf_再構成レポート.xlsx")
    wb.save(save_path)
    print(f"Excel created at: {save_path}")

if __name__ == "__main__":
    create_reconstructed_excel()
