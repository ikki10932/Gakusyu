import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_summary_excel():
    wb = openpyxl.Workbook()
    
    header_font = Font(name='Meiryo UI', bold=True, color="FFFFFF")
    body_font = Font(name='Meiryo UI')
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_header_style(sheet, columns):
        for col_idx, (name, width) in enumerate(columns, 1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        sheet.freeze_panes = "A4"

    # Sheet 1: レビュー結果概要
    ws1 = wb.active
    ws1.title = "レビュー結果概要"
    cols1 = [("項目", 25), ("内容", 80)]
    apply_header_style(ws1, cols1)
    
    data1 = [
        ("対象タイトル", "AIエージェントによるソースコード解析と資料生成の有効性検証レポート"),
        ("改善の方向性", "「試行錯誤の履歴」から「確立された手法の共有・第三者が再現できる手順の体系化」へシフトする。"),
        ("エグゼクティブサマリー", "・プロセス記録から再現可能な「ワークフロー手順書」への移行\n・特定AIモデル依存記述の削減と指示思想の汎用化\n・ローカルパス役割の明示と依存関係の可視化")
    ]
    
    for r_idx, row in enumerate(data1, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = top_left_alignment
            cell.border = border

    # Sheet 2: ディスカッション・改善項目
    ws2 = wb.create_sheet("改善項目詳細")
    cols2 = [("カテゴリ", 20), ("課題・項目", 35), ("詳細説明", 60), ("対応状況", 15)]
    apply_header_style(ws2, cols2)
    
    data2 = [
        ("問題点", "前提と結果の混在", "目的・コマンド・所感が1つのセクションに混在し結論が特定困難", ""),
        ("問題点", "モデル固有への過度な依存", "特定のAIモデルのトライアル状況が目立ち、モデル代替わり時の陳腐化リスク", ""),
        ("問題点", "参照パスの不透明性", "ローカルパスの役割説明が不足し、環境変更時に追跡不能になる", ""),
        ("構成要件", "1. 検証の背景と目的", "属人化解消の課題とAI適用の狙いを明文化", ""),
        ("構成要件", "2. 検証環境とツールスタック", "Antigravity、使用対象とするSKILLの役割定義", ""),
        ("構成要件", "3. 標準解析ワークフロー", "特定のモデルに依存しない、入力から出力を得るまでの汎用的な手順", ""),
        ("構成要件", "4. 検証結果とベストプラクティス", "出力物の評価と、プロンプトの工夫点を明記", ""),
        ("構成要件", "5. 発生した課題と対処法", "エラー発生時の機械的なリトライ手順等のルール化", ""),
        ("構成要件", "6. リファレンス", "使用した設定ファイル、対象ソースの概要、用語定義", ""),
        ("補強提案", "トラブルシューティングの標準化", "「Agent terminated...」発生時等、定量的・具体的な対処手順をまとめる", ""),
        ("補強提案", "指示書(SKILL等)の管理方針", "AIへの指示をコード同様に構成管理し、意図や変更履歴を残す運用提案", ""),
        ("欠落情報", "核となるポリシー・制約事項", "rules.mdやInstructions.mdなどの具体的コンテンツ", ""),
        ("欠落情報", "定量的な効果測定結果", "AIを使わなかった場合との工数比較など", ""),
        ("欠落情報", "解析対象関数の選定理由", "プロンプトで当該関数を指定した背景", "")
    ]
    
    for r_idx, row in enumerate(data2, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = center_alignment if c_idx in [1, 4] else top_left_alignment
            cell.border = border

    output_dir = r"d:\_Develop\Skills\MySkills\review\output"
    os.makedirs(output_dir, exist_ok=True)
    save_path = os.path.join(output_dir, "2026-03-02_1pdf_資料設計レビュー_テンプレート.xlsx")
    wb.save(save_path)
    print(f"Excel created at: {save_path}")

if __name__ == "__main__":
    create_summary_excel()
