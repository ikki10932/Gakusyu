import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_summary_excel():
    wb = openpyxl.Workbook()
    
    header_font = Font(name='Meiryo UI', bold=True, color="FFFFFF")
    body_font = Font(name='Meiryo UI')
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_header_style(sheet, columns):
        for col_idx, (name, width) in enumerate(columns, 1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        sheet.freeze_panes = "A4"

    # Sheet 1: 概要とサマリー
    ws1 = wb.active
    ws1.title = "概要とサマリー"
    cols1 = [("項目", 25), ("内容", 80)]
    apply_header_style(ws1, cols1)
    
    data1 = [
        ("タイトル", "AIエージェントによるソースコード解析と資料生成の有効性検証レポート"),
        ("概要", "Antigravity等のAIを活用し、ソースコード解析および資料作成業務において、長期的な保守性と理解容易性を確保するための設計手法と検証結果のまとめ。"),
        ("改善の方向性", "「検証プロセス（試行錯誤）」から「知見の共有と手法の標準化」へシフトし、第三者が再現できる形式へ整理する。"),
        ("エグゼクティブサマリー", "・目的はAIによる解析業務の標準化と品質向上\n・現状のプロセス主眼から再現可能なワークフロー提示へ移行すべき\n・モデル依存記述を減らし、指示(SKILL)の設計思想の明文化が不可欠")
    ]
    
    for r_idx, row in enumerate(data1, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = top_left_alignment
            cell.border = border

    # Sheet 2: レビュー詳細と推奨構成
    ws2 = wb.create_sheet("レビュー詳細と推奨構成")
    cols2 = [("カテゴリ", 20), ("項目", 35), ("詳細・理由", 60)]
    apply_header_style(ws2, cols2)
    
    data2 = [
        ("問題点", "情報の混在", "目的、コマンド、結果が1つのセクションに混在し結論が判別しにくい"),
        ("問題点", "コンテキストの依存", "特定のAIモデル(Gemini3.1pro等)の挙動記述に依存し、数年後の有用性が低下する懸念"),
        ("問題点", "リンクの不透明性", "多数のローカルパスについて役割説明がなく、環境変化時に追跡不能になるリスク"),
        ("推奨構成", "1. 検証の背景と目的", "現状の課題とAI活用による解決目標"),
        ("推奨構成", "2. 検証環境と使用ツール", "Antigravityバージョン、使用したSKILL.mdの役割定義"),
        ("推奨構成", "3. 標準解析ワークフロー", "モデル非依存の汎用的なプロンプトと実行手順"),
        ("推奨構成", "4. 検証結果と考察", "生成物例、モデル特性、エラー対処法"),
        ("推奨構成", "5. 運用留意事項とベストプラクティス", "SKILL.md改善点、バッチ処理化指針"),
        ("推奨構成", "6. 付録：リファレンス", "設定ファイル群、出力格納先、用語定義"),
        ("必須記載事項", "標準解析ワークフロー", "入力データから出力を得るまでの論理的ステップ"),
        ("必須記載事項", "検証結果", "生成結果が「長期的に理解可能」かどうかの評価"),
        ("補強提案", "エラーハンドリング集", "「Agent terminated」等の頻発エラーへの具体的・機械的な対処手順のリスト化"),
        ("補強提案", "SKILL.mdのバージョン管理", "解析ロジックをコード管理し、変更履歴を追跡可能にする方針"),
        ("欠落情報", "核となる指示内容", "rules.mdやInstructions.mdの具体的なポリシー"),
        ("欠落情報", "AI未使用時との比較データ", "定量的な効果測定データ")
    ]
    
    for r_idx, row in enumerate(data2, 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = center_alignment if c_idx == 1 else top_left_alignment
            cell.border = border

    output_dir = r"d:\_Develop\Skills\MySkills\review\output"
    os.makedirs(output_dir, exist_ok=True)
    save_path = os.path.join(output_dir, "2026-03-02_資料レビューまとめ.xlsx")
    wb.save(save_path)
    print(f"Excel created at: {save_path}")

if __name__ == "__main__":
    create_summary_excel()
